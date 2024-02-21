from typing import List

from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.workbook import Workbook

from beans.TableBean import TableBean
from driver.xlsx.XlsxService import XlsxService
from driver.xlsx.template.NHCDBXlsxTemplate import NHCDBXlsxTemplate


class XlsxServiceImpl(XlsxService):
    # 工作book
    workbook: Workbook = None
    template: NHCDBXlsxTemplate = None

    ali = Alignment(horizontal='center', vertical='center')
    # 内容字体颜色
    font = Font(name='Arial', size=8, color='000000')
    # 第二行标题颜色
    title_font = Font(name='Arial', size=10, color='000000', bold=True)
    # 第一行字体颜色
    first_row_font = Font(name='Arial', size=15, color='000000', bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    def __init__(self, workbook: Workbook):
        super().__init__()
        self.workbook = workbook

    def insert_tables(self, sheet_name: str, table_beans: List[TableBean]) -> None:
        if table_beans is None:
            return

        row_offset = 0
        for table_bean in table_beans:
            self.insert_table(sheet_name, row_offset, table_bean)
            row_offset = row_offset + len(table_bean.table_columns) + 2

    def insert_table(self, sheet_name: str, row_offset: int, table_bean: TableBean) -> None:
        """在excel的某个sheet中插入一个表格"""
        # 插入第一行，并且合并单元格
        if sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
        else:
            worksheet = self.workbook.create_sheet(sheet_name)
        # 添加第一行表头
        end_column = len(NHCDBXlsxTemplate.title)
        end_row = len(table_bean.table_columns) + 2 + row_offset
        worksheet.merge_cells(start_row=1 + row_offset, end_row=1 + row_offset, start_column=1, end_column=end_column)

        table_title = table_bean.table_name + '-' + (table_bean.table_comment or "")
        merge_cell = worksheet.cell(row=1 + row_offset, column=1, value=table_title)
        # merge_cell.font = self.first_row_font
        # 添加第二行表头
        worksheet.append(NHCDBXlsxTemplate.title)

        table_column_list = NHCDBXlsxTemplate.parse_table_column(table_bean.table_columns)

        for table_column in table_column_list:
            worksheet.append(table_column)

        # todo 设置列宽 改成动态的
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 10
        worksheet.column_dimensions['G'].width = 10
        worksheet.column_dimensions['H'].width = 10
        worksheet.column_dimensions['I'].width = 10
        worksheet.column_dimensions['J'].width = 10
        worksheet.column_dimensions['K'].width = 10
        worksheet.column_dimensions['L'].width = 20

        # 设置行高

        worksheet.row_dimensions[1 + row_offset].height = 25
        worksheet.row_dimensions[2 + row_offset].height = 20
        for index, value in enumerate(table_column_list):
            worksheet.row_dimensions[index + row_offset + 3].height = 25

        # 设置边框样式
        for row in worksheet.iter_rows(min_row=1 + row_offset, max_row=end_row, min_col=1, max_col=end_column):
            for cell in row:
                cell.border = self.border
                cell.font = self.font

        for row in worksheet.iter_rows(min_row=2 + row_offset, max_row=2 + row_offset, min_col=1, max_col=end_column):
            for cell in row:
                cell.font = self.title_font

        for row in worksheet.iter_rows(min_row=1 + row_offset, max_row=1 + row_offset, min_col=1, max_col=end_column):
            for cell in row:
                cell.font = self.first_row_font

        # 设置全部单元格居中显示
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = self.ali
