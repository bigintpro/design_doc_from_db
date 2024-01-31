from openpyxl import Workbook
# Python types will automatically be converted
import datetime

if __name__ == '__main__':
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")