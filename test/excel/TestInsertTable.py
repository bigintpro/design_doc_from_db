from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

if __name__ == '__main__':
    wb: Workbook = Workbook()
    ws: Worksheet = wb.active

    ws['A1'] = "TPMSXD"
    ws.merge_cells('A1:L1')
    ws.append(["序号", "字段名称", "字段注释", "数据类型", "字段长度", "小数位数", "值域", "默认值", "约束", "键",
               "非空性", "备注"])

    ws.append(["1", "ID", "ID", "VARCHAR2", "32", "", "", "", "主键约束", "PK", "TRUE", ""])
    table = Table(name="abc", displayName="Table1", ref="A2:L3")
    ws.add_table(table)
    wb.save("insertTable.xlsx")
