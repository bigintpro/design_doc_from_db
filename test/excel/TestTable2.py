from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ali = Alignment(horizontal='center', vertical='center')

    # 添加表头
    ws.merge_cells("A1:L1")

    ws["A1"] = "WS"
    ws['A1'].alignment = ali  # 居中显示

    ws.append(["序号", "字段名称", "字段注释", "数据类型", "字段长度", "小数位数", "值域", "默认值", "约束", "键",
               "非空性", "备注"])

    ws.append(["1", "ID", "ID", "VARCHAR2", "32", "", "", "", "主键约束", "PK", "TRUE", ""])
    ws.append(["1", "ID", "ID", "VARCHAR2", "32", "", "", "", "主键约束", "PK", "TRUE", ""])

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 35
    ws.row_dimensions[4].height = 35

    font = Font(name='Arial', size=8, color='000000')

    # 设置边框样式
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=12):
        for cell in row:
            cell.border = border
            cell.font = font

    # 设置全部单元格居中显示
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = ali

    # ws.auto_filter.ref = None
    wb.save('table2.xlsx')
